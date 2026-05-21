#!/usr/bin/env python3
"""Collect Amazon listing images, A+ images, and SellerSprite reviews for one ASIN."""

from __future__ import annotations

import argparse
import datetime as dt
import html
import json
import os
import re
import sys
import time
import urllib.error
import urllib.request
import uuid
from pathlib import Path
from urllib.parse import urlparse
from typing import Any

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError:
    Workbook = None


HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 Chrome/120 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}
MCP_URL = "https://mcp.sellersprite.com/mcp"
KEY_ENV_NAMES = (
    "SELLERSPRITE_SECRET_KEY",
    "SELLERSPRITE_API_KEY",
    "SELLERSPRITE_MCP_SECRET_KEY",
)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Create a Desktop ASIN package with main images, A+ images, and reviews Excel.")
    parser.add_argument("input", help="Amazon ASIN or product page URL.")
    parser.add_argument("--marketplace", default="US", help="SellerSprite marketplace code. Default: US.")
    parser.add_argument("--out", help="Output directory. Default: Desktop/<ASIN>.")
    parser.add_argument("--main-count", type=int, help="Limit main image count.")
    parser.add_argument("--max-aplus-images", type=int, default=40, help="Maximum A+ image count. Default: 40.")
    parser.add_argument("--skip-aplus", action="store_true", help="Do not download A+ images.")
    parser.add_argument("--skip-reviews", action="store_true", help="Do not export SellerSprite reviews.")
    parser.add_argument("--max-review-pages", type=int, default=10000, help="Maximum review pages to fetch. Default: 10000.")
    parser.add_argument("--review-sleep", type=float, default=1.6, help="Delay between SellerSprite review pages. Default: 1.6 seconds.")
    parser.add_argument("--secret-key", help="SellerSprite secret key. Prefer environment variables.")
    parser.add_argument("--mcp-url", default=MCP_URL, help=f"SellerSprite MCP endpoint. Default: {MCP_URL}")
    parser.add_argument("--save-links", action="store_true", help="Save image URL link lists.")
    args = parser.parse_args()
    if args.max_review_pages < 1:
        parser.error("--max-review-pages must be >= 1")
    if args.max_aplus_images < 1:
        parser.error("--max-aplus-images must be >= 1")
    if args.review_sleep < 0:
        parser.error("--review-sleep must be >= 0")
    return args


def asin_from_input(value: str) -> str:
    value = value.strip()
    if re.fullmatch(r"[A-Z0-9]{10}", value.upper()):
        return value.upper()
    patterns = [
        r"/dp/([A-Z0-9]{10})",
        r"/gp/aw/d/([A-Z0-9]{10})",
        r"/gp/product/([A-Z0-9]{10})",
        r"[/&?]pd_rd_i=([A-Z0-9]{10})",
    ]
    for pattern in patterns:
        match = re.search(pattern, value)
        if match:
            return match.group(1).upper()
    raise SystemExit("Could not identify ASIN from input.")


def fetch_bytes(url: str) -> bytes:
    request = urllib.request.Request(url, headers=HEADERS)
    with urllib.request.urlopen(request, timeout=60) as response:
        return response.read()


def fetch_text(url: str) -> str:
    return fetch_bytes(url).decode("utf-8", errors="replace")


def clean_text(value: Any) -> Any:
    if not isinstance(value, str):
        return value
    value = value.replace("<br>", "\n").replace("<br/>", "\n").replace("<br />", "\n")
    value = re.sub(r"<[^>]+>", "", value)
    return html.unescape(value).strip()


def extract_balanced_json_array(page: str, marker: str) -> list[Any]:
    start = page.find(marker)
    if start < 0:
        return []
    pos = start + len(marker)
    if pos >= len(page) or page[pos] != "[":
        return []
    depth = 0
    in_str = False
    esc = False
    end = None
    for index, char in enumerate(page[pos:], start=pos):
        if in_str:
            if esc:
                esc = False
            elif char == "\\":
                esc = True
            elif char == '"':
                in_str = False
        else:
            if char == '"':
                in_str = True
            elif char == "[":
                depth += 1
            elif char == "]":
                depth -= 1
                if depth == 0:
                    end = index + 1
                    break
    if end is None:
        return []
    try:
        parsed = json.loads(page[pos:end])
    except json.JSONDecodeError:
        return []
    return parsed if isinstance(parsed, list) else []


def extract_main_image_urls(page: str) -> list[str]:
    gallery = extract_balanced_json_array(page, "'colorImages': { 'initial': ")
    urls: list[str] = []
    for item in gallery:
        if not isinstance(item, dict):
            continue
        url = item.get("hiRes") or item.get("large")
        if url and "m.media-amazon.com/images/I/" in url and url not in urls:
            urls.append(url.replace("\\u002F", "/").replace("\\/", "/"))
    if urls:
        return urls
    for pattern in [r'"hiRes":"(https:[^"]+)"', r'"large":"(https:[^"]+)"', r'data-old-hires="([^"]+)"']:
        for match in re.finditer(pattern, page):
            url = match.group(1).replace("\\u002F", "/").replace("\\/", "/")
            if "m.media-amazon.com/images/I/" in url and url not in urls:
                urls.append(url)
    return urls


def extract_aplus_html(page: str) -> str:
    starts = [
        index for index in [
            page.find('<div id="aplus"'),
            page.find("<div id='aplus'"),
            page.find('id="aplus_feature_div"'),
            page.find("id='aplus_feature_div'"),
        ] if index >= 0
    ]
    if not starts:
        return ""
    start = min(starts)
    stop_terms = [
        'id="productDetails_feature_div"',
        "id='productDetails_feature_div'",
        'id="detailBullets',
        "id='detailBullets",
        'id="customerReviews"',
        "id='customerReviews'",
        'id="reviewsMedley"',
        "id='reviewsMedley'",
    ]
    stops = [page.find(term, start + 1) for term in stop_terms]
    stops = [index for index in stops if index > start]
    end = min(stops) if stops else min(len(page), start + 350000)
    return page[start:end]


def normalize_amazon_image_url(url: str) -> str:
    url = html.unescape(url).replace("\\u002F", "/").replace("\\/", "/")
    return re.sub(r"\._[A-Z0-9_,]+_\.", ".", url)


def extract_aplus_image_urls(page: str, main_urls: list[str], limit: int) -> list[str]:
    region = extract_aplus_html(page)
    if not region:
        return []
    urls: list[str] = []
    for pattern in [r'<img[^>]+(?:src|data-src)="([^"]+)"', r'"(https://m\.media-amazon\.com/images/[^"]+)"']:
        for match in re.finditer(pattern, region, flags=re.I):
            url = normalize_amazon_image_url(match.group(1))
            if "m.media-amazon.com/images/" not in url:
                continue
            if any(url == main or url in main or main in url for main in main_urls):
                continue
            if url not in urls:
                urls.append(url)
                if len(urls) >= limit:
                    return urls
    return urls


def download_images(urls: list[str], out_dir: Path) -> list[dict[str, Any]]:
    out_dir.mkdir(parents=True, exist_ok=True)
    rows = []
    for index, url in enumerate(urls, start=1):
        suffix = Path(urlparse(url).path).suffix or ".jpg"
        if suffix.lower() not in {".jpg", ".jpeg", ".png", ".webp"}:
            suffix = ".jpg"
        path = out_dir / f"{index:02d}{suffix}"
        data = fetch_bytes(url)
        path.write_bytes(data)
        rows.append({"index": index, "file": path.name, "url": url, "bytes": len(data)})
    return rows


def get_secret_key(explicit: str | None) -> str | None:
    if explicit:
        return explicit
    for name in KEY_ENV_NAMES:
        value = os.environ.get(name)
        if value:
            return value
    return None


def post_json(url: str, secret_key: str, payload: dict[str, Any]) -> dict[str, Any]:
    body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    request = urllib.request.Request(
        url,
        data=body,
        method="POST",
        headers={
            "secret-key": secret_key,
            "Content-Type": "application/json;charset=utf-8",
            "Accept": "application/json, text/event-stream",
            "x-request-id": str(uuid.uuid4()),
            "User-Agent": "codex-amazon-images-reviews/1.0",
        },
    )
    with urllib.request.urlopen(request, timeout=60) as response:
        return json.loads(response.read().decode("utf-8"))


def call_review_tool(secret_key: str, mcp_url: str, payload: dict[str, Any], request_id: int) -> dict[str, Any]:
    rpc_payload = {
        "jsonrpc": "2.0",
        "id": request_id,
        "method": "tools/call",
        "params": {"name": "review", "arguments": payload},
    }
    response = post_json(mcp_url, secret_key, rpc_payload)
    if "error" in response:
        error = response["error"]
        raise RuntimeError(f"MCP error: {error.get('code')} {error.get('message')}")
    content = ((response.get("result") or {}).get("content") or [])
    if not content or not content[0].get("text"):
        raise RuntimeError("MCP response did not include review JSON.")
    return json.loads(content[0]["text"])


def iso_date(ms: Any) -> Any:
    if not isinstance(ms, int):
        return ms
    return dt.datetime.fromtimestamp(ms / 1000, tz=dt.timezone.utc).date().isoformat()


def normalize_review(item: dict[str, Any]) -> dict[str, Any]:
    normalized = dict(item)
    if "content" in normalized:
        normalized["content"] = clean_text(normalized["content"])
    if "date" in normalized:
        normalized["dateIso"] = iso_date(normalized["date"])
    return normalized


def fetch_all_reviews(args: argparse.Namespace, asin: str, secret_key: str) -> dict[str, Any]:
    all_items: list[dict[str, Any]] = []
    first_data: dict[str, Any] | None = None
    for page in range(1, args.max_review_pages + 1):
        response = call_review_tool(
            secret_key,
            args.mcp_url,
            {"marketplace": args.marketplace, "asin": asin, "page": page, "size": 10},
            page,
        )
        if response.get("code") != "OK":
            raise RuntimeError(f"SellerSprite review error: {response.get('code')} {response.get('message', '')}")
        data = response.get("data") or {}
        if first_data is None:
            first_data = data
        items = [normalize_review(item) for item in data.get("items") or []]
        all_items.extend(items)
        total_pages = data.get("pages")
        reached_last_page = isinstance(total_pages, int) and page >= total_pages
        if not items or data.get("hasNextPage") is False or reached_last_page:
            break
        if args.review_sleep:
            time.sleep(args.review_sleep)
    return {
        "metadata": {
            "asin": asin,
            "marketplace": args.marketplace,
            "fetchedItems": len(all_items),
            "apiTotal": first_data.get("total") if first_data else None,
            "apiPages": first_data.get("pages") if first_data else None,
            "provider": "sellersprite-mcp",
        },
        "items": all_items,
    }


def write_reviews_excel(path: Path, result: dict[str, Any]) -> None:
    if Workbook is None:
        raise RuntimeError("openpyxl is required to write Excel. Use the Codex bundled Python runtime or install openpyxl.")
    wb = Workbook()
    ws = wb.active
    ws.title = "Reviews"
    headers = [
        "No", "ASIN", "Marketplace", "Date", "Star", "Title", "Content", "Author",
        "Verified Purchase", "Vine", "Has Image", "Has Video", "Likes", "SKUs",
        "Author Labels", "Image URLs", "Video URLs",
    ]
    ws.append(headers)
    meta = result["metadata"]
    for index, item in enumerate(result["items"], start=1):
        ws.append([
            index,
            meta.get("asin"),
            meta.get("marketplace"),
            item.get("dateIso") or item.get("date"),
            item.get("star"),
            item.get("title"),
            item.get("content"),
            item.get("author"),
            item.get("verified"),
            item.get("vine"),
            item.get("image"),
            item.get("video"),
            item.get("likes"),
            "\n".join(item.get("skus") or []),
            "\n".join(item.get("authorLabels") or []),
            "\n".join(item.get("images") or []),
            "\n".join(item.get("videos") or []),
        ])
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    widths = {
        "A": 8, "B": 14, "C": 12, "D": 14, "E": 8, "F": 38, "G": 80,
        "H": 24, "I": 18, "J": 10, "K": 12, "L": 12, "M": 10,
        "N": 35, "O": 22, "P": 70, "Q": 50,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    summary = wb.create_sheet("Summary")
    for row in [
        ("ASIN", meta.get("asin")),
        ("Marketplace", meta.get("marketplace")),
        ("Fetched Items", meta.get("fetchedItems")),
        ("API Total", meta.get("apiTotal")),
        ("API Pages", meta.get("apiPages")),
        ("Provider", meta.get("provider")),
    ]:
        summary.append(row)
    for cell in summary["A"]:
        cell.font = Font(bold=True)
    summary.column_dimensions["A"].width = 18
    summary.column_dimensions["B"].width = 24
    wb.save(path)


def write_links(path: Path, rows: list[dict[str, Any]]) -> None:
    path.write_text("\n".join(f"{row['index']:02d}. {row['url']}" for row in rows), encoding="utf-8")


def main() -> int:
    args = parse_args()
    asin = asin_from_input(args.input)
    product_url = f"https://www.amazon.com/dp/{asin}?th=1"
    out_dir = Path(args.out) if args.out else Path.home() / "Desktop" / asin
    out_dir.mkdir(parents=True, exist_ok=True)

    page = fetch_text(product_url)
    main_urls = extract_main_image_urls(page)
    if args.main_count is not None:
        main_urls = main_urls[: args.main_count]
    if not main_urls:
        raise SystemExit("No current-variant main gallery images found.")
    main_rows = download_images(main_urls, out_dir / "main-images")

    aplus_rows: list[dict[str, Any]] = []
    if not args.skip_aplus:
        aplus_urls = extract_aplus_image_urls(page, main_urls, args.max_aplus_images)
        if aplus_urls:
            aplus_rows = download_images(aplus_urls, out_dir / "aplus-images")

    review_path = None
    review_error = None
    review_count = 0
    if not args.skip_reviews:
        secret_key = get_secret_key(args.secret_key)
        if not secret_key:
            review_error = "Missing SellerSprite secret key. Set SELLERSPRITE_SECRET_KEY or pass --secret-key."
        else:
            try:
                reviews = fetch_all_reviews(args, asin, secret_key)
                review_count = len(reviews["items"])
                review_path = out_dir / f"{asin}-reviews.xlsx"
                write_reviews_excel(review_path, reviews)
            except Exception as exc:
                review_error = str(exc)

    if args.save_links:
        write_links(out_dir / f"{asin}-main-image-urls.txt", main_rows)
        if aplus_rows:
            write_links(out_dir / f"{asin}-aplus-image-urls.txt", aplus_rows)

    manifest = {
        "asin": asin,
        "productUrl": product_url,
        "folder": str(out_dir),
        "mainImages": main_rows,
        "aplusImages": aplus_rows,
        "reviewsExcel": str(review_path) if review_path else None,
        "reviewCount": review_count,
        "reviewError": review_error,
    }
    manifest_path = out_dir / f"{asin}-manifest.json"
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(manifest, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    sys.exit(main())
